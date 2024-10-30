from openpyxl import load_workbook
import pandas as pd


def coordinate_data():
    coordinate_excel = load_workbook(filename="../EnvParams/Layout_Coordinates.xlsx", data_only=True)
    coordinate_excel = coordinate_excel.active
    coordinate_sheet = list(coordinate_excel.values)[1:]

    coordinate_data = {}
    for coordinate in coordinate_sheet:
        coordinate_data[coordinate[0]] = (coordinate[1], coordinate[2])
    return coordinate_data


def all_hunger_data():
    hunger_excel = load_workbook(filename="../Beachmark/dynamic/16.xlsx", data_only=True)
    sheet_names = hunger_excel.sheetnames
    sheet_count = len(sheet_names)
    all_hunger_data = []
    for _ in range(sheet_count):
        one_hunger_data = {}
        hunger_sheet = list(hunger_excel.worksheets[_].values)[1:]
        for one_hunger in hunger_sheet:
            if one_hunger[0] is not None:
                one_hunger_data[one_hunger[0]] = one_hunger[1]
        all_hunger_data.append(one_hunger_data)
    return all_hunger_data


def coefficient_data():
    feeding_adjustment_coefficient_excel = load_workbook(filename="../EnvParams/Feeding_Coefficient.xlsx", data_only=True)
    feeding_adjustment_coefficient_excel = feeding_adjustment_coefficient_excel.active
    feeding_adjustment_coefficient_sheet = list(feeding_adjustment_coefficient_excel.values)[1:]

    coefficient_data = {}
    for feeding_adjustment_coefficient in feeding_adjustment_coefficient_sheet:
        coefficient_data[feeding_adjustment_coefficient[0]] = feeding_adjustment_coefficient[1]
    return coefficient_data


def available_trolley_data():
    available_trolley_excel = load_workbook(filename='../EnvParams/Available_Trolley.xlsx')
    available_trolley_sheet = available_trolley_excel.active

    available_trolley_data = {}
    for row in available_trolley_sheet.iter_rows(min_row=2, values_only=True):
        available_trolley = [available_trolley for idx, available_trolley in enumerate(row[1:]) if available_trolley is not None]
        available_trolley_data[row[0]] = tuple(available_trolley)
    return available_trolley_data


def scheduling_data_to_excel(path, data):
    headline = ['trolley number', 'destination', 'begin time', 'end time', 'remaining load']
    df = pd.DataFrame(data, columns=headline)
    df.to_excel(path, index=False)


def fitness_data_to_excel(path, data):
    df = pd.DataFrame(data)
    df.to_excel(path, index=False)


coordinate_data = coordinate_data()
all_hunger_data = all_hunger_data()
scheduling_num = len(all_hunger_data)
coefficient_data = coefficient_data()
available_trolley_data = available_trolley_data()
trolley_num = 6
supply_depot_num = 2
hunger_threshold = 50
w = 1.4
l = 3.6
h = 2
UW = 30
LW = 2
SP = 1
SF = 0.1
SL = 0.5

if __name__ == "__main__":
    print(all_hunger_data)

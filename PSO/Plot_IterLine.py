from NSGA.Gantt import line_chart

from openpyxl import load_workbook
import matplotlib.pyplot as plt


def scheduling_data(filename):
    scheduling_excel = load_workbook(filename=filename, data_only=True)
    scheduling_excel = scheduling_excel.active
    scheduling_sheet = list(scheduling_excel.values)[1:]
    scheduling_data = []
    for scheduling in scheduling_sheet:
        scheduling_data.append(scheduling)
    return scheduling_data


def fitness_data(filename):
    fitness_excel = load_workbook(filename=filename, data_only=True)
    fitness_excel = fitness_excel.active
    fitness_sheet = list(fitness_excel.values)[1:]
    fitness_data = []
    for fitness in fitness_sheet:
        fitness_data.append(fitness)
    return fitness_data


if __name__ == "__main__":
    seed_num = 1
    scheduling_cout = 1
    file1 = 'pso'
    file2 = 'nsga'
    file3 = 'ensga'
    beachmark = 'dynamic16'

    # fitness line
    filename1 = f'../FitnessData/{beachmark}/{file1}/scheduling{str(scheduling_cout)}_{str(seed_num)}.xlsx'
    fitness_data1 = fitness_data(filename1)
    # filename2 = f'../FitnessData/{beachmark}/{file2}/scheduling{scheduling_cout}_{str(seed_num)}.xlsx'
    # fitness_data2 = fitness_data(filename2)
    # filename3 = f'../FitnessData/{beachmark}/{file3}/scheduling{scheduling_cout}_{str(seed_num)}.xlsx'
    # fitness_data3 = fitness_data(filename3)

    for idx, fitness in enumerate(fitness_data1):
        label = 'PSO(f' + str(idx+1) + ')'
        line_chart(label, fitness, 300)
    # for idx, fitness in enumerate(fitness_data2):
    #     label = 'NSGA-Ⅱ(f' + str(idx+1) + ')'
    #     line_chart(label, fitness, 300)
    # for idx, fitness in enumerate(fitness_data3):
    #     label = 'ENSGA-Ⅱ(f' + str(idx + 1) + ')'
    #     line_chart(label, fitness, 300)
    plt.legend(loc="upper left", bbox_to_anchor=(1, 1))
    plt.show()

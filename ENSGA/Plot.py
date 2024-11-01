from Data import scheduling_num
from Gantt import gantt, line_chart
from Iteration import scheduling_interval

from openpyxl import load_workbook
import matplotlib.pyplot as plt


def scheduling_data(filename):
    scheduling_excel = load_workbook(filename=filename, data_only=True)
    scheduling_excel = scheduling_excel.active
    scheduling_sheet = list(scheduling_excel.values)[1:]
    scheduling_data = []
    for scheduling in scheduling_sheet:
        scheduling_data.append(scheduling)
    return scheduling_data


def fitness_data(filename):
    fitness_excel = load_workbook(filename=filename, data_only=True)
    fitness_excel = fitness_excel.active
    fitness_sheet = list(fitness_excel.values)[1:]
    fitness_data = []
    for fitness in fitness_sheet:
        fitness_data.append(fitness)
    return fitness_data


if __name__ == "__main__":
    seed_num = 1
    scheduling_cout = 1
    file = 'ensga'
    # scheduling gantt
    filename = f'../SchedulingData/dynamic9/{file}/scheduling{str(scheduling_cout)}_{str(seed_num)}.xlsx'
    scheduling_data = scheduling_data(filename)
    if scheduling_cout != scheduling_num:
        schedule_stime = scheduling_cout*scheduling_interval
    else:
        schedule_stime = float('inf')
    gantt(1, schedule_stime, scheduling_data)
    plt.show()

    # fitness line
    # filename = '../FitnessData/static1' + file + '/scheduling' + str(scheduling_cout) + str(seed_num) + '.xlsx'
    # fitness_data1 = fitness_data(filename)
    # for fit_label, fitness in enumerate(fitness_data1):
    #     line_chart(fit_label, fitness, 300)
    # plt.show()

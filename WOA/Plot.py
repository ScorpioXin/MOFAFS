from NSGA.Gantt import line_chart

from openpyxl import load_workbook
import matplotlib.pyplot as plt


def scheduling_data(filename):
    scheduling_excel = load_workbook(filename=filename, data_only=True)
    scheduling_excel = scheduling_excel.active
    scheduling_sheet = list(scheduling_excel.values)[1:]
    scheduling_data = []
    for scheduling in scheduling_sheet:
        scheduling_data.append(scheduling)
    return scheduling_data


def fitness_data(filename):
    fitness_excel = load_workbook(filename=filename, data_only=True)
    fitness_excel = fitness_excel.active
    fitness_sheet = list(fitness_excel.values)[1:]
    fitness_data = []
    for fitness in fitness_sheet:
        fitness_data.append(fitness)
    return fitness_data


if __name__ == "__main__":
    seed_num = 1
    scheduling_cout = 1
    file0 = 'gwo'
    file1 = 'woa'
    file2 = 'pso'
    file3 = 'nsga'
    file4 = 'ensga'
    beachmark = 'static3'

    # fitness line
    # filename0 = f'../FitnessData/{beachmark}/{file0}/scheduling{scheduling_cout}_{seed_num}.xlsx'
    # fitness_data0 = fitness_data(filename0)
    filename1 = f'../FitnessData/{beachmark}/{file1}/scheduling{scheduling_cout}_{seed_num}.xlsx'
    fitness_data1 = fitness_data(filename1)
    # filename2 = f'../FitnessData/{beachmark}/{file2}/scheduling{scheduling_cout}_{seed_num}.xlsx'
    # fitness_data2 = fitness_data(filename2)
    # filename3 = f'../FitnessData/{beachmark}/{file3}/scheduling{scheduling_cout}_{seed_num}.xlsx'
    # fitness_data3 = fitness_data(filename3)
    # filename4 = f'../FitnessData/{beachmark}/{file4}/scheduling{scheduling_cout}_{seed_num}.xlsx'
    # fitness_data4 = fitness_data(filename4)

    if scheduling_cout == 1:
        workbook = load_workbook(f'../FitnessData/{beachmark}/first_init_fitness.xlsx')
        sheet = workbook.active
    elif scheduling_cout == 2:
        workbook = load_workbook(f'../FitnessData/{beachmark}/second_init_fitness.xlsx')
        sheet = workbook.active

    # for idx, fitness in enumerate(fitness_data0):
    #     init_fit = sheet.cell(row=idx+2, column=seed_num).value
    #     fitness = list(fitness)
    #     fitness.insert(0, init_fit)
    #     label = 'GWO(f' + str(idx+1) + ')'
    #     line_chart(label, fitness, 300)
    for idx, fitness in enumerate(fitness_data1):
        init_fit = sheet.cell(row=idx+2, column=seed_num).value
        fitness = list(fitness)
        fitness.insert(0, init_fit)
        label = 'WOA(f' + str(idx + 1) + ')'
        line_chart(label, fitness, 300)
    # for idx, fitness in enumerate(fitness_data2):
    #     init_fit = sheet.cell(row=idx+2, column=seed_num).value
    #     fitness = list(fitness)
    #     fitness.insert(0, init_fit)
    #     label = 'PSO(f' + str(idx+1) + ')'
    #     line_chart(label, fitness, 300)
    # for idx, fitness in enumerate(fitness_data3):
    #     init_fit = sheet.cell(row=idx+2, column=seed_num).value
    #     fitness = list(fitness)
    #     fitness.insert(0, init_fit)
    #     label = 'NSGA-Ⅱ(f' + str(idx+1) + ')'
    #     line_chart(label, fitness, 300)
    # for idx, fitness in enumerate(fitness_data4):
    #     init_fit = sheet.cell(row=idx+2, column=seed_num).value
    #     fitness = list(fitness)
    #     fitness.insert(0, init_fit)
    #     label = 'ENSGA-Ⅱ(f' + str(idx + 1) + ')'
    #     line_chart(label, fitness, 300)
    plt.legend(loc="upper left", bbox_to_anchor=(1, 1), prop={'size': 15})
    plt.show()

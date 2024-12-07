from Gantt_Line import gantt

from openpyxl import load_workbook
import matplotlib.pyplot as plt


if __name__ == "__main__":
    beachmark = 'dynamic10'
    scheduling_interval = 200
    run_cout = 1

    all_scheduling_list = []
    workbook1 = load_workbook(f'SchedulingData/{beachmark}/ensga/scheduling1_{run_cout}.xlsx')
    sheet1 = workbook1.active
    scheduling_list = []
    for row in sheet1.iter_rows(min_row=2, max_col=5, values_only=True):
        scheduling_list.append(list(row))
    all_scheduling_list.append(scheduling_list)

    if "dynamic" in beachmark:
        workbook2 = load_workbook(f'SchedulingData/{beachmark}/ensga/scheduling2_{run_cout}.xlsx')
        sheet2 = workbook2.active
        scheduling_list = []
        for row in sheet2.iter_rows(min_row=2, max_col=5, values_only=True):
            scheduling_list.append(list(row))
        all_scheduling_list.append(scheduling_list)

    for figure_num, scheduling_list in enumerate(all_scheduling_list):
        if figure_num != len(all_scheduling_list)-1:
            schedule_stime = (figure_num+1)*scheduling_interval
        else:
            schedule_stime = float('inf')
        gantt(figure_num, schedule_stime, scheduling_list)
    plt.show()

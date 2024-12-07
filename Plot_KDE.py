from openpyxl import load_workbook
import matplotlib.pyplot as plt
import matplotlib
import seaborn as sns
matplotlib.use('TkAgg')


if __name__ == "__main__":
    scheduling_count = 2
    start, middle, end = 1, 150, 300
    beachmark = "dynamic15"
    # workbook = load_workbook(f'SolutionDistribution/{beachmark}_{start}.xlsx')
    workbook = load_workbook(f'SolutionDistribution/{beachmark}_{scheduling_count}_{start}.xlsx')
    sheet = workbook.active
    fitness_list1 = [cell.value for cell in sheet[2]]
    fitness_list2 = [cell.value for cell in sheet[3]]
    plt.figure(figsize=(10, 6))
    plt.rcParams['font.sans-serif'] = ['Times New Roman']
    plt.rcParams['font.size'] = 35
    sns.kdeplot(x=fitness_list1, y=fitness_list2, cmap="YlOrBr", fill=True, thresh=0, levels=20, alpha=0.6)
    plt.scatter(fitness_list1, fitness_list2, color='red', alpha=0.4, s=100)
    plt.title("Distribution of fitness at the beginning of the iteration")
    plt.xlabel("Weighted makespan (f1)")
    plt.ylabel("Maximum load of AFVs (f2)")

    # workbook = load_workbook(f'SolutionDistribution/{beachmark}_{middle}.xlsx')
    workbook = load_workbook(f'SolutionDistribution/{beachmark}_{scheduling_count}_{middle}.xlsx')
    sheet = workbook.active
    fitness_list3 = [cell.value for cell in sheet[2]]
    fitness_list4 = [cell.value for cell in sheet[3]]
    plt.figure(figsize=(10, 6))
    sns.kdeplot(x=fitness_list3, y=fitness_list4, cmap="Greens", fill=True, thresh=0, levels=20, alpha=0.6)
    plt.scatter(fitness_list3, fitness_list4, color='red', alpha=0.4, s=100)
    plt.title("Distribution of fitness at mid-iteration")
    plt.xlabel("Weighted makespan (f1)")
    plt.ylabel("Maximum load of AFVs (f2)")

    # workbook = load_workbook(f'SolutionDistribution/{beachmark}_{end}.xlsx')
    workbook = load_workbook(f'SolutionDistribution/{beachmark}_{scheduling_count}_{end}.xlsx')
    sheet = workbook.active
    fitness_list5 = [cell.value for cell in sheet[2]]
    fitness_list6 = [cell.value for cell in sheet[3]]
    plt.figure(figsize=(10, 6))
    sns.kdeplot(x=fitness_list5, y=fitness_list6, cmap="Blues", fill=True, thresh=0, levels=20, alpha=0.6)
    plt.scatter(fitness_list5, fitness_list6, color='red', alpha=0.4, s=100)
    plt.title("Distribution of fitness at the end of the iteration")
    plt.xlabel("Weighted makespan (f1)")
    plt.ylabel("Maximum load of AFVs (f2)")

    plt.show()

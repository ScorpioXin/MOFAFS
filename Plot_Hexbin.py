from openpyxl import load_workbook
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('TkAgg')


scheduling_count = 1
start, middle, end = 1, 150, 300
workbook = load_workbook(f'SolutionDistribution/static5_{start}.xlsx')
# workbook = load_workbook(f'SolutionDistribution/dynamic15_{scheduling_count}_{start}.xlsx')
sheet = workbook.active
fitness_list1 = [cell.value for cell in sheet[2]]
fitness_list2 = [cell.value for cell in sheet[3]]
plt.figure(figsize=(10, 6))
plt.hexbin(fitness_list1, fitness_list2, gridsize=30, cmap='Blues', mincnt=1)
cb = plt.colorbar(label='Counts')
plt.title("Distribution of fitness at the beginning of the iteration")
plt.xlabel("Weighted makespan (f1)")
plt.ylabel("Maximum load of AFVs (f2)")
plt.grid(True)

workbook = load_workbook(f'SolutionDistribution/static5_{middle}.xlsx')
# workbook = load_workbook(f'SolutionDistribution/dynamic15_{scheduling_count}_{middle}.xlsx')
sheet = workbook.active
fitness_list3 = [cell.value for cell in sheet[2]]
fitness_list4 = [cell.value for cell in sheet[3]]
plt.figure(figsize=(10, 6))
plt.hexbin(fitness_list3, fitness_list4, gridsize=30, cmap='Greens', mincnt=1)
cb = plt.colorbar(label='Counts')
plt.title("Distribution of fitness at mid-iteration")
plt.xlabel("Weighted makespan (f1)")
plt.ylabel("Maximum load of AFVs (f2)")
plt.grid(True)

workbook = load_workbook(f'SolutionDistribution/static5_{end}.xlsx')
# workbook = load_workbook(f'SolutionDistribution/dynamic15_{scheduling_count}_{end}.xlsx')
sheet = workbook.active
fitness_list5 = [cell.value for cell in sheet[2]]
fitness_list6 = [cell.value for cell in sheet[3]]
plt.figure(figsize=(10, 6))
plt.hexbin(fitness_list5, fitness_list6, gridsize=30, cmap='Reds', mincnt=1)
cb = plt.colorbar(label='Counts')
plt.title("Distribution of fitness at the end of the iteration")
plt.xlabel("Weighted makespan (f1)")
plt.ylabel("Maximum load of AFVs (f2)")
plt.grid(True)

plt.show()
